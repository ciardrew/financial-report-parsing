import pandas as pd
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import graph_creation as gc

section_headers = ["Income", "Expenditure", "Operating Surplus / Deficit before Contributions and Depn", "Intra-Regional Contribution", "Surplus / Deficit"]
months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

def xlsx_append(df_dict, date, complete_path):
    """Appends data to an existing Excel file with formatting."""
    workbook = load_workbook(complete_path)
    sheetnames = [s for s in workbook.sheetnames if s != "Data Visualization"]
    split_sheetnames = [s.split(" ")[0] for s in sheetnames]
    print(sheetnames)
    print(date)
    print(date.split(" ")[0])
    if date not in sheetnames or date.split(" ")[0] not in split_sheetnames:
        sheet = workbook.create_sheet(title=date)
        row_offset = 1

        # Styles
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        bold_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # headers
        sheet.cell(row=row_offset, column=1, value="Order of St John").fill = orange_fill
        sheet.cell(row=row_offset, column=1).font = bold_font
        sheet.cell(row=row_offset, column=1).border = border

        sheet.cell(row=row_offset + 1, column=1, value="Income and Expenditure by Cost Centre").fill = orange_fill
        sheet.cell(row=row_offset + 1, column=1).font = bold_font
        sheet.cell(row=row_offset + 1, column=1).border = border

        sheet.cell(row=row_offset + 2, column=1, value=f"{date}").fill = orange_fill
        sheet.cell(row=row_offset + 2, column=1).font = bold_font
        sheet.cell(row=row_offset + 2, column=1).border = border

        row_offset += 5

        for branch, df in df_dict.items():
            # Headers with orange format
            for col_num, col_name in enumerate(df.columns, start=1):
                if col_num == 1:
                    sheet.cell(row=row_offset, column=1, value=branch).fill = orange_fill
                else:
                    sheet.cell(row=row_offset, column=col_num, value=col_name).fill = orange_fill
                sheet.cell(row=row_offset, column=col_num).font = bold_font
                sheet.cell(row=row_offset, column=col_num).border = border
            row_offset += 1

            # descriptions
            for _, row in df.iterrows():
                if row['description'] in section_headers:
                    for col_num, value in enumerate(row, start=1):
                        cell = sheet.cell(row=row_offset, column=col_num, value=value)
                        cell.font = bold_font
                        cell.border = border
                    row_offset += 1
                else:
                    for col_num, value in enumerate(row, start=1):
                        cell = sheet.cell(row=row_offset, column=col_num, value=value)
                        if col_num == 1 and not row['is_bold']:
                            cell.alignment = Alignment(indent=4)
                row_offset += 1

            # Adjust column widths based on contents
            for col_num, column_name in enumerate(df.columns, start=1):
                column_len = max(df[column_name].astype(str).map(len).max(), len(column_name))
                sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = column_len
                
            row_offset += 4  # Add 3 rows for a gap before the next table

        workbook.save(complete_path)

        gc.graph_sheet_creation(complete_path)  # Call the graph creation function to add graphs to the sheet
    else:
        print("Sheet already created for that report")