import report_append as ra
import data_processing as dp
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

def is_cost_centre(name):
    return isinstance(name, str) and name.startswith('SA') and ' - ' in name

def graph_sheet_creation(complete_path):
    workbook = load_workbook(complete_path)
    sheetnames = [s for s in workbook.sheetnames if s != "Data Visualization"]
    
    # Dictionary to hold cost centre data
    cost_centre_data = {}

    for month in sheetnames:
        ws = workbook[month]
        current_centre = None
        for row in ws.iter_rows(min_row=4, values_only=True):
            description = row[0]
            actual_month = row[1]

            if is_cost_centre(description):
                current_centre = description
                if current_centre not in cost_centre_data:
                    cost_centre_data[current_centre] = {}
                if month not in cost_centre_data[current_centre]:
                    cost_centre_data[current_centre][month] = {'income': None, 'expenditure': None}
            elif current_centre: # only if currcentre exists
                if description == "Income":
                    cost_centre_data[current_centre][month]['income'] = actual_month
                    #print(f"{current_centre} on {month}: has {actual_month} income")
                elif description == "Expenditure":
                    cost_centre_data[current_centre][month]['expenditure'] = actual_month
                    #print(f"{current_centre} on {month}: has {actual_month} expenditure")

    if "Data Visualization" in workbook.sheetnames:
        vis_sheet = workbook["Data Visualization"]
        vis_sheet._images = []
    else:
        vis_sheet = workbook.create_sheet(title="Data Visualization")

    img_row = 2
    for centre, month_dict in cost_centre_data.items():
        months = []
        incomes = []
        expenditures = []
        for month, values in month_dict.items():
            months.append(month)
            incomes.append(values['income'] if values['income'] is not None else 0)
            expenditures.append(values['expenditure'] if values['expenditure'] is not None else 0)
        
        months_shorthand = [month.split(" ")[0][:3] + month.split(" ")[1][2:] for month in months]

        plt.figure(figsize=(8, 5))
        plt.plot(months_shorthand, incomes, label='Income', linestyle='solid')
        plt.plot(months_shorthand, expenditures, label='Expenditure', linestyle='solid')
        plt.xlabel('Month')
        plt.ylabel('Amount (NZD)')
        plt.title(f'{centre} - Income/Expenditure/Total by Month')
        plt.legend()
        plt.tight_layout()
        img_filename = f"{centre.replace(' ', '_').replace('/', '_')}_plot.png"
        plt.savefig(img_filename)
        plt.close()

        img = Image(img_filename)
        vis_sheet.add_image(img, f'B{img_row}')
        img_row += 30 
        
    workbook.save(complete_path)